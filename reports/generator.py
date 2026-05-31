import pandas as pd
import json
from datetime import datetime
from typing import List
from config import REPORTS_DIR
from utils.logger import logger
from utils.schema import FinalReportRecord

def generate_reports(records: List[FinalReportRecord], base_filename: str = "audit_report"):
    """
    Generates CSV, Excel, and JSON reports from the final verified records.
    """
    if not records:
        logger.warning("No records provided to generate reports.")
        return

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    prefix = f"{base_filename}_{timestamp}"
    
    # Convert Pydantic models to dicts
    data = [r.model_dump() for r in records]
    df = pd.DataFrame(data)

    # Filter and rename columns to exactly match user's requested layout
    display_columns = {
        'row_number': 'S.No',
        'institute_name': 'Institute Name',
        'course_name': 'Course Name',
        'verification_status': 'Match Status',
        'confidence_score': 'Confidence',
        'verified_institute_name': 'Verified Institute',
        'original_mode': 'Original Mode',
        'verified_mode': 'Verified Mode',
        'original_country': 'Original Country',
        'verified_country': 'Verified Country',
        'original_skills': 'Original Skills',
        'verified_skills': 'Verified Skills',
        'original_fees': 'Original Fees',
        'verified_fees': 'Verified Fees',
        'original_logo': 'PDF Logo Present',
        'verified_logo': 'Verified Logo',
        'ai_summary': 'AI Summary',
        'course_link': 'Course Link',
        'response_time_ms': 'Response Time (ms)',
        'screenshot_path': 'Error Screenshot'
    }
    
    # Ensure all required columns exist before filtering to avoid KeyError
    available_cols = [col for col in display_columns.keys() if col in df.columns]
    df_display = df[available_cols].rename(columns=display_columns)

    # File paths
    csv_path = REPORTS_DIR / f"{prefix}.csv"
    excel_path = REPORTS_DIR / f"{prefix}.xlsx"
    json_path = REPORTS_DIR / f"{prefix}.json"

    logger.info("Generating reports...")

    try:
        # Save CSV
        df_display.to_csv(csv_path, index=False)
        logger.info(f"Saved CSV report: {csv_path}")

        # Save Excel with column width auto-adjustment and conditional formatting
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_display.to_excel(writer, index=False, sheet_name='Audit_Report')
            worksheet = writer.sheets['Audit_Report']
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Beautiful Styles
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill("solid", fgColor="2C3E50") # Enterprise Dark Blue
            wrap_alignment = Alignment(wrap_text=True, vertical="top")
            thin_border = Border(bottom=Side(style='thin', color='E0E0E0'))
            
            # Status colors
            color_map = {
                'VALID': PatternFill("solid", fgColor="C6EFCE"), # Green
                'PARTIAL_MATCH': PatternFill("solid", fgColor="FFEB9C"), # Yellow
                'INVALID': PatternFill("solid", fgColor="FFC7CE"), # Light Red
                'BROKEN_LINK': PatternFill("solid", fgColor="9C0006") # Dark Red
            }
            
            # Format Headers
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Freeze Top Row
            worksheet.freeze_panes = "A2"
            
            # Find status column index
            status_col_idx = None
            for idx, col in enumerate(df_display.columns):
                if col == 'Match Status':
                    status_col_idx = idx + 1
                    break
            
            # Format Columns & Rows
            for idx, col in enumerate(df_display.columns):
                col_letter = get_column_letter(idx + 1)
                
                # Set Breathable Column Widths
                if col == 'AI Summary':
                    worksheet.column_dimensions[col_letter].width = 60
                elif col in ['Course Name', 'Institute Name', 'Course Link']:
                    worksheet.column_dimensions[col_letter].width = 45
                elif col == 'S.No':
                    worksheet.column_dimensions[col_letter].width = 8
                elif 'Skills' in col:
                    worksheet.column_dimensions[col_letter].width = 30
                elif 'Verified' in col or 'Original' in col:
                    worksheet.column_dimensions[col_letter].width = 20
                else:
                    worksheet.column_dimensions[col_letter].width = 18
                
                # Apply wrapping and borders to all cells
                for cell in worksheet[col_letter]:
                    if cell.row != 1:
                        cell.alignment = wrap_alignment
                        cell.border = thin_border
                        # Apply conditional coloring based on the Match Status row
                        if status_col_idx:
                            status_val = worksheet.cell(row=cell.row, column=status_col_idx).value
                            if status_val in color_map:
                                # Apply color just to the status cell for cleanliness, or the whole row
                                if cell.column == status_col_idx:
                                    cell.fill = color_map[status_val]
                                    if status_val == 'BROKEN_LINK':
                                        cell.font = Font(color="FFFFFF", bold=True)
                                        
        logger.info(f"Saved Excel report: {excel_path}")

        # Save JSON
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4)
        logger.info(f"Saved JSON report: {json_path}")

    except Exception as e:
        logger.error(f"Failed to generate reports: {e}")
